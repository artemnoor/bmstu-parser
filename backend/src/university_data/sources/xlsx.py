from __future__ import annotations

from collections.abc import Iterator
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


class XlsxExtractor:
    """Read workbook values without requiring Excel or LibreOffice."""

    def rows(self, path: Path, *, sheet: str | None = None) -> Iterator[dict[str, Any]]:
        workbook = load_workbook(path, read_only=False, data_only=True)
        try:
            worksheet = workbook[sheet] if sheet else workbook.active
            merged_values: dict[tuple[int, int], Any] = {}
            for merged in worksheet.merged_cells.ranges:
                anchor = worksheet.cell(merged.min_row, merged.min_col).value
                for row in range(merged.min_row, merged.max_row + 1):
                    for column in range(merged.min_col, merged.max_col + 1):
                        merged_values[(row, column)] = anchor
            values = list(worksheet.iter_rows(values_only=True))
            if not values:
                return
            headers = [str(value or "").strip() for value in values[0]]
            for row_number, values_row in enumerate(values[1:], start=2):
                result: dict[str, Any] = {"_row": row_number}
                for column, header in enumerate(headers, start=1):
                    if not header:
                        continue
                    value = (
                        values_row[column - 1] if column <= len(values_row) else None
                    )
                    if value in (None, "") and (row_number, column) in merged_values:
                        value = merged_values[(row_number, column)]
                    result[header] = value
                yield result
        finally:
            workbook.close()
